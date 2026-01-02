
from __future__ import annotations
from pathlib import Path
import argparse, sys, math, glob, os
from xlsx_loader import load_groupe_tp, load_groups
from greedy import creer_groupes_glouton
from rules import score_equilibre, validite_globale

DEFAULT_XLSX_CANDIDATES = [
    "Groupes SAÉ S3 -constitution.xlsx",
    "Groupes SAE S3 -constitution.xlsx",
    "groupes.xlsx",
]

def _auto_pick_xlsx(explicit: str | None) -> str:
    if explicit and Path(explicit).exists():
        return explicit
    # try known names
    for name in DEFAULT_XLSX_CANDIDATES:
        if Path(name).exists():
            return name
    # fallback: first .xlsx in cwd
    files = sorted(glob.glob("*.xlsx"))
    if files:
        return files[0]
    raise FileNotFoundError("Aucun fichier .xlsx trouvé dans le dossier courant — passez --xlsx.")

def _auto_pick_sheet(wb_path: str, explicit: str | None) -> str:
    if explicit:
        return explicit
    try:
        # prefer "Liste S3" if present, else first sheet
        from openpyxl import load_workbook
        wb = load_workbook(filename=wb_path, read_only=True)
        if "Liste S3" in wb.sheetnames:
            return "Liste S3"
        return wb.sheetnames[0]
    except Exception:
        return "Liste S3"

def afficher_repartition(rep) -> None:
    print("=== Répartition des groupes ===")
    for nom, gr in sorted(rep.groups.items()):
        print(f"\n{nom} (cap {gr.room}, total avantage={gr.avantage_total():.2f})")
        for e in gr.members:
            tag = "⭐" if e.leader else " "
            pol = f"[p{e.polarite}]" if e.polarite is not None else ""
            print(f"  {tag} {e.prenom} {e.nom:<20} {pol:>5}  adv={e.avantage:g}")
    print(f"\nScore d'équilibre (max-min) = {score_equilibre(rep):.2f}")
    print("Valide (leaders & polarités):", "✅" if validite_globale(rep) else "❌")

def main():
    p = argparse.ArgumentParser(description="Création de groupes projet (glouton) — usage simple : python cli.py <GROUPE_TP>")
    # Positional: nom du groupe TP (ex: 1A, G1-TPA, etc.)
    p.add_argument("groupe", help="Nom du groupe TP tel qu'indiqué dans la colonne 'Groupe' du XLSX (ex: 1A)")

    # Optionnels (avec auto-détection et valeurs par défaut)
    p.add_argument("--xlsx", help="Chemin du fichier Excel (.xlsx). Si omis, tentative d'auto-détection dans le dossier.")
    p.add_argument("--feuille", help="Nom de la feuille (par défaut: 'Liste S3' si existe, sinon première feuille).")
    p.add_argument("--nb-groupes", type=int, default=0, help="Nombre de groupes projet à créer. Si 0, calculé automatiquement via --taille-cible.")
    p.add_argument("--taille-cible", type=int, default=4, help="Taille cible par groupe (utilisée si --nb-groupes=0).")
    p.add_argument("--seed", type=int, default=None, help="(Optionnel) graine aléatoire pour départager les égalités.")

    args = p.parse_args()

    # Résoudre xlsx et feuille automatiquement si besoin
    xlsx = _auto_pick_xlsx(args.xlsx)
    sheet = _auto_pick_sheet(xlsx, args.feuille)

    # Charger le groupe TP demandé
    try:
        gtp = load_groupe_tp(xlsx, sheet, args.groupe)
    except KeyError as e:
        # proposer les groupes disponibles pour aider
        try:
            all_groups = sorted(load_groups(xlsx, sheet).keys())
            hint = f"Groupes disponibles dans '{sheet}': " + (', '.join(all_groups[:30]) + (' ...' if len(all_groups) > 30 else ''))
        except Exception:
            hint = "Impossible de lister les groupes disponibles (problème de lecture XLSX)."
        raise SystemExit(f"Groupe '{args.groupe}' introuvable dans {xlsx}/{sheet}.\n{hint}") from e

    # Calcul du nombre de groupes si non fourni
    nb_groupes = args.nb_groupes if args.nb_groupes and args.nb_groupes > 0 else max(1, math.ceil(len(gtp.etudiants) / max(1, args.taille_cible)))

    rep = creer_groupes_glouton(gtp, nb_groupes, seed=args.seed)
    afficher_repartition(rep)

if __name__ == "__main__":
    main()
